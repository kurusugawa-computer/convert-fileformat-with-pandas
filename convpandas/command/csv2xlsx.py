import argparse
import collections
import sys
import warnings
from pathlib import Path
from typing import Dict, List, TextIO, Union

import openpyxl
import pandas

from convpandas.common.cli import PrettyHelpFormatter

MAXIMUM_NUMBER_OF_CHARACTERS_OF_SHEET_NAME = 31


def _read_csv(
    csv_file: Union[str, TextIO],
    sep: str,
    encoding: str,
    quotechar: str,
) -> pandas.DataFrame:
    read_csv_kwargs = {
        "sep": sep,
        "encoding": encoding,
        "quotechar": quotechar,
        "header": None,
    }
    return pandas.read_csv(csv_file, **read_csv_kwargs)


def _to_numeric(value):
    return pandas.to_numeric(value, errors="ignore")


def _do_not_anything(value):
    return value


def _to_excel(
    df_dict: Dict[str, pandas.DataFrame],
    xlsx_file: str,
    string_to_numeric: bool = True,
) -> None:
    Path(xlsx_file).parent.mkdir(exist_ok=True, parents=True)
    if string_to_numeric:
        convert_func = _to_numeric
    else:
        convert_func = _do_not_anything

    workbook = openpyxl.Workbook(write_only=True)
    for sheet_name, df in df_dict.items():
        if sheet_name == "-":
            worksheet = workbook.create_sheet()
        else:
            if len(sheet_name) > MAXIMUM_NUMBER_OF_CHARACTERS_OF_SHEET_NAME:
                warnings.warn(
                    f"Sheet name '{sheet_name}' is more than 31 characters. So sheet name is truncated."
                )
                sheet_name = sheet_name[0:MAXIMUM_NUMBER_OF_CHARACTERS_OF_SHEET_NAME]
            worksheet = workbook.create_sheet(title=sheet_name)
        values = df.values
        for row_index in range(df.shape[0]):
            row = values[row_index]
            worksheet.append([convert_func(value) for value in row])
    workbook.save(xlsx_file)


def csv2xlsx(
    csv_files: List[str],
    xlsx_file: str,
    *,
    sep: str,
    encoding: str,
    quotechar: str,
    string_to_numeric: bool,
    sheet_names: List[str],
):
    if sheet_names is not None:
        if len(sheet_names) != len(csv_files):
            print(
                "Size of csv_files and size of sheet_name do not match.",
                file=sys.stderr,
            )
            sys.exit(1)

    df_dict: Dict[str, pandas.DataFrame] = collections.OrderedDict()

    if len(csv_files) == 1 and csv_files[0] == "-":
        df = _read_csv(sys.stdin, sep=sep, encoding=encoding, quotechar=quotechar)
        if sheet_names is None:
            df_dict["-"] = df
        else:
            df_dict[sheet_names[0]] = df
        _to_excel(df_dict, xlsx_file=xlsx_file, string_to_numeric=string_to_numeric)
        return

    if sheet_names is None:
        for file in csv_files:
            file_path = Path(file)
            if not file_path.exists():
                print(f"No such file: '{file}'", file=sys.stderr)
                sys.exit(1)

            df = _read_csv(file, sep=sep, encoding=encoding, quotechar=quotechar)
            df_dict[file_path.stem] = df

    else:
        for file, name in zip(csv_files, sheet_names):
            file_path = Path(file)
            if not file_path.exists():
                print(f"No such file: '{file}'", file=sys.stderr)
                sys.exit(1)

            df = _read_csv(file, sep=sep, encoding=encoding, quotechar=quotechar)
            df_dict[name] = df

    _to_excel(df_dict, xlsx_file=xlsx_file, string_to_numeric=string_to_numeric)


def main(args):
    csv2xlsx(
        csv_files=args.csv_files,
        xlsx_file=args.xlsx_file,
        sep=args.sep,
        encoding=args.encoding,
        quotechar=args.quotechar,
        string_to_numeric=not args.numeric_to_string,
        sheet_names=args.sheet_name,
    )


def add_parser(subparsers: argparse._SubParsersAction):
    parser = subparsers.add_parser(
        "csv2xlsx",
        help="Convert csv file to xlsx file.",
        formatter_class=PrettyHelpFormatter,
    )
    parser.set_defaults(command_help=parser.print_help)

    parser.add_argument("csv_files", type=str, nargs="+")
    parser.add_argument("xlsx_file", type=str)

    parser.add_argument("--sep", default=",", help="Delimiter to use when reading csv.")

    parser.add_argument(
        "--encoding",
        default="utf-8",
        help="Encoding to use when reading csv. List of Python standard encodings.\n"
        "https://docs.python.org/3/library/codecs.html#standard-encodings",
    )
    parser.add_argument(
        "--quotechar",
        default='"',
        help="The character used to denote the start and end of a quoted item when reading csv.",
    )

    parser.add_argument(
        "--numeric_to_string",
        action="store_true",
        help="If specified, write numeric value as string type. If not specified, write numeric value as numeric type.",
    )

    parser.add_argument("--sheet_name", type=str, nargs="+")

    parser.set_defaults(subcommand_func=main)
    return parser
