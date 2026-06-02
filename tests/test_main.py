import io
import os
import sys
from pathlib import Path

import openpyxl
import pandas

from convpandas.__main__ import cli
from convpandas.command.csv2xlsx import _to_excel

# プロジェクトトップに移動する
os.chdir(os.path.dirname(os.path.abspath(__file__)) + "/../")

out_path = Path("./tests/out/")
out_path.mkdir(exist_ok=True, parents=True)

data_path = Path("./tests/data")


class Test_csv2xlsx:
    def test_standard(self):  # noqa: ANN201
        cli(["csv2xlsx", str(data_path / "test.csv"), str(out_path / "out.xlsx")])

    def test_convert_stdin_csv(self):  # noqa: ANN201
        sys.stdin = io.StringIO("name,id\nAlice,1\n")
        cli(["csv2xlsx", "-", str(out_path / "out2.xlsx")])

    def test_convert_multiple_csv_to_xlsx(self):  # noqa: ANN201
        cli(
            [
                "csv2xlsx",
                str(data_path / "test.csv"),
                str(data_path / "test2.csv"),
                str(data_path / "test.csv"),
                str(out_path / "out3.xlsx"),
            ]
        )

    def test_convert_multiple_csv_to_xlsx_with_sheetnames(self):  # noqa: ANN201
        cli(
            [
                "csv2xlsx",
                str(data_path / "test.csv"),
                str(data_path / "test2.csv"),
                str(data_path / "test.csv"),
                str(out_path / "out3.xlsx"),
                "--sheet_name",
                "alice bob",
                "bob",
                "chris",
            ]
        )

    def test_write_pandas3_string_dtype_missing_value_as_blank_cell(self, tmp_path):  # noqa: ANN001, ANN201
        df = pandas.DataFrame({0: pandas.Series(["name", None], dtype="str")})
        xlsx_file = tmp_path / "out.xlsx"

        _to_excel({"sheet": df}, str(xlsx_file), string_to_numeric=False)

        worksheet = openpyxl.load_workbook(xlsx_file).active
        assert worksheet["A1"].value == "name"
        assert worksheet["A2"].value is None

    def test_write_nullable_string_dtype_missing_value_as_blank_cell(self, tmp_path):  # noqa: ANN001, ANN201
        df = pandas.DataFrame({0: pandas.Series(["name", pandas.NA], dtype="string")})
        xlsx_file = tmp_path / "out.xlsx"

        _to_excel({"sheet": df}, str(xlsx_file), string_to_numeric=False)

        worksheet = openpyxl.load_workbook(xlsx_file).active
        assert worksheet["A1"].value == "name"
        assert worksheet["A2"].value is None

    def test_convert_numeric_string_dtype_values_to_excel_numbers(self, tmp_path):  # noqa: ANN001, ANN201
        df = pandas.DataFrame({0: pandas.Series(["001", None], dtype="str")})
        xlsx_file = tmp_path / "out.xlsx"

        _to_excel({"sheet": df}, str(xlsx_file), string_to_numeric=True)

        worksheet = openpyxl.load_workbook(xlsx_file).active
        assert worksheet["A1"].value == 1
        assert worksheet["A2"].value is None


class Test_xlsx2csv:
    def test_standard(self):  # noqa: ANN201
        cli(["xlsx2csv", str(data_path / "test.xlsx"), str(out_path / "out.csv")])

    def test_specify_sheetname(self):  # noqa: ANN201
        cli(
            [
                "xlsx2csv",
                str(data_path / "test.xlsx"),
                str(out_path / "out2.csv"),
                "--sheet_name",
                "bob",
            ]
        )

    def test_output_stdout_csv(self, capsys):  # noqa: ANN001, ANN201
        cli(["xlsx2csv", str(data_path / "test2.xlsx"), "-"])

        captured = capsys.readouterr()
        assert captured.out == "name,age\n田中,23\n"
